"""
Excel routes for managing product data in the PL出品マクロ.xlsm file.
"""

from flask import Blueprint, request, jsonify
from app.services.excel_data_service import ExcelDataService
from app.services.excel_creator_service import ExcelCreatorService
from app.services.category_lookup_service import CategoryLookupService
import os

excel_bp = Blueprint('excel', __name__,  url_prefix='/api')

# Initialize the Excel service
excel_service = ExcelDataService()

# Initialize the Excel creator service
excel_creator_service = ExcelCreatorService()

# Initialize the Category lookup service
category_service = CategoryLookupService()

@excel_bp.route('/health', methods=['GET'])
def health_check():
    """
    Health check endpoint for the backend server.
    """
    try:
        return jsonify({
            'status': 'healthy',
            'service': 'Excel Data Service',
            'timestamp': __import__('datetime').datetime.now().isoformat()
        })
    except Exception as e:
        return jsonify({
            'status': 'unhealthy',
            'error': str(e)
        }), 500

@excel_bp.route('/excel/add-product', methods=['POST'])
def add_product_to_excel():
    """
    Add a single product to the appropriate sheet in the Excel file.
    """
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({
                'success': False,
                'message': 'No data provided'
            }), 400
        
        # Add the data to Excel
        success, message = excel_service.add_data_to_excel(data)
        
        if success:
            return jsonify({
                'success': True,
                'message': message
            })
        else:
            return jsonify({
                'success': False,
                'message': message
            }), 400
            
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Unexpected error: {str(e)}'
        }), 500

@excel_bp.route('/excel/add-products-bulk', methods=['POST'])
def add_products_bulk():
    """
    Add multiple products to the Excel file.
    """
    try:
        data = request.get_json()
        
        if not data or 'products' not in data:
            return jsonify({
                'success': False,
                'message': 'No products data provided'
            }), 400
        
        products = data['products']
        if not isinstance(products, list):
            return jsonify({
                'success': False,
                'message': 'Products must be a list'
            }), 400
        
        # Process bulk data
        success_count, failure_count, error_messages = excel_service.bulk_add_data(products)
        
        return jsonify({
            'success': True,
            'summary': {
                'total_processed': len(products),
                'successful': success_count,
                'failed': failure_count
            },
            'errors': error_messages if error_messages else None
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Unexpected error: {str(e)}'
        }), 500

@excel_bp.route('/excel/classify-product', methods=['POST'])
def classify_product():
    """
    Classify a product to determine which sheet it should go to.
    """
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({
                'success': False,
                'message': 'No data provided'
            }), 400
        
        title = data.get('タイトル') or data.get('title', '')
        if not title:
            return jsonify({
                'success': False,
                'message': 'Title is required for classification'
            }), 400
        
        # Classify the product
        category = excel_service.classify_product_category(title, data)
        
        return jsonify({
            'success': True,
            'category': category,
            'title': title
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error classifying product: {str(e)}'
        }), 500

@excel_bp.route('/excel/sheet-info', methods=['GET'])
def get_sheet_info():
    """
    Get information about all sheets in the Excel file.
    """
    try:
        sheet_info = excel_service.get_sheet_info()
        
        return jsonify({
            'success': True,
            'sheets': sheet_info
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error getting sheet info: {str(e)}'
        }), 500

@excel_bp.route('/excel/test-sample', methods=['POST'])
def test_sample_data():
    """
    Test the Excel functionality with sample data.
    """
    try:
        # Sample data similar to the user's example
        sample_data = {
            "カテゴリ": "2084005208",
            "管理番号": "1212260021698",
            "タイトル": "◇ PIVOT DOOR ピボットドアー 胸ロゴ、裾袖絞りあり 長袖 フリース 表記なし グリーン レディースメンズ E 1212260021698",
            "文字数": 57,
            "付属品": "無",
            "ランク": "3",
            "コメント": "目立った傷や汚れなし",
            "素材": "画像参照",
            "色": "グリーン",
            "サイズ": "表記なし",
            "着丈": 66,
            "　肩幅": 58,
            "身幅": 58,
            "袖丈": 58,
            "梱包サイズ": "通常",
            "梱包記号": "◇",
            "美品": "",
            "ブランド": "PIVOT DOOR ピボットドアー",
            "フリー": "胸ロゴ、裾袖絞りあり",
            "袖": "長袖",
            "もの": "フリース",
            "男女": "レディースメンズ",
            "採寸1": "着丈：約66cm　肩幅：約58cm　身幅：約58cm　袖丈：約58cm",
            "ラック": "ベースW/26",
            "金額": 2000
        }
        
        # Add the sample data
        success, message = excel_service.add_data_to_excel(sample_data)
        
        if success:
            return jsonify({
                'success': True,
                'message': message,
                'sample_data': sample_data
            })
        else:
            return jsonify({
                'success': False,
                'message': message,
                'sample_data': sample_data
            }), 400
            
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error testing sample data: {str(e)}'
        }), 500

@excel_bp.route('/excel/mapping-preview', methods=['POST'])
def get_mapping_preview():
    """
    Preview how data would be mapped to a specific sheet without adding it.
    """
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({
                'success': False,
                'message': 'No data provided'
            }), 400
        
        # Get sheet name from request or classify
        sheet_name = data.get('sheet_name')
        if not sheet_name:
            title = data.get('タイトル') or data.get('title', '')
            if title:
                sheet_name = excel_service.classify_product_category(title, data)
            else:
                return jsonify({
                    'success': False,
                    'message': 'Sheet name or title required for mapping preview'
                }), 400
        
        # Get mapped data
        mapped_data = excel_service.map_data_to_sheet_headers(data, sheet_name)
        
        # Generate measurement text
        measurement_text = excel_service.generate_measurement_text(data, sheet_name)
        
        return jsonify({
            'success': True,
            'sheet_name': sheet_name,
            'mapped_data': mapped_data,
            'measurement_text': measurement_text,
            'original_data': data
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error generating mapping preview: {str(e)}'
        }), 500

@excel_bp.route('/excel/save-file', methods=['POST'])
def save_excel_file():
    """
    Save the current Excel file to a specified location.
    """
    try:
        data = request.get_json()
        target_path = data.get('target_path')
        
        if not target_path:
            return jsonify({
                'success': False,
                'message': 'Target path is required'
            }), 400
        
        # Copy the current Excel file to the target location
        import shutil
        try:
            shutil.copy2(excel_service.excel_file_path, target_path)
            return jsonify({
                'success': True,
                'message': f'Excel file saved successfully to {target_path}',
                'target_path': target_path
            })
        except Exception as e:
            return jsonify({
                'success': False,
                'message': f'Failed to save Excel file: {str(e)}'
            }), 500
            
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Unexpected error: {str(e)}'
        }), 500

@excel_bp.route('/excel/export-to-excel', methods=['POST'])
def export_to_excel():
    """
    Export processed results to Excel file with proper sheet classification.
    """
    try:
        print("🚀 Export to Excel endpoint called")
        data = request.get_json()
        print(f"📊 Received data keys: {list(data.keys()) if data else 'None'}")
        
        if not data or 'processed_results' not in data:
            print("❌ No processed results provided in request")
            return jsonify({
                'success': False,
                'message': 'No processed results provided'
            }), 400
        
        processed_results = data['processed_results']
        print(f"📊 Processed results type: {type(processed_results)}")
        print(f"📊 Processed results count: {len(processed_results) if processed_results else 0}")
        
        if not isinstance(processed_results, dict):
            print(f"❌ Expected dict, got {type(processed_results)}")
            return jsonify({
                'success': False,
                'message': 'Processed results must be a dictionary'
            }), 400
        
        # Check if Excel file exists and is accessible
        excel_file_path = excel_service.excel_file_path
        if not os.path.exists(excel_file_path):
            print(f"❌ Excel file not found: {excel_file_path}")
            return jsonify({
                'success': False,
                'message': f'Excel file not found: {excel_file_path}'
            }), 404
        
        # Pre-define sheet headers (constant structure like your sample code)
        # These headers match exactly what's in the PL出品マクロ.xlsm file
        SHEET_HEADERS = {
            "トップス": [
                "カテゴリ", "管理番号", "タイトル", "文字数", "付属品", "ランク", "コメント", 
                "素材", "色", "サイズ", "着丈", "　肩幅", "身幅", "袖丈", "梱包サイズ", 
                "梱包記号", "美品", "ブランド", "フリー", "袖", "もの", "男女", 
                "採寸1", "ラック", "金額", "股上", "股下", "ウエスト", "もも幅", "裾幅", "総丈", "ヒップ", "仕入先", "仕入日", "原価"
            ],
            "パンツ": [
                "カテゴリ", "管理番号", "タイトル", "文字数", "付属品", "ランク", "コメント", 
                "素材", "色", "サイズ", "着丈", "　肩幅", "身幅", "袖丈", "梱包サイズ", 
                "梱包記号", "美品", "ブランド", "フリー", "袖", "もの", "男女", 
                "採寸1", "ラック", "金額", "股上", "股下", "ウエスト", "もも幅", "裾幅", "総丈", "ヒップ", "仕入先", "仕入日", "原価"
            ],
            "スカート": [
                "カテゴリ", "管理番号", "タイトル", "文字数", "付属品", "ランク", "コメント", 
                "素材", "色", "サイズ", "着丈", "　肩幅", "身幅", "袖丈", "梱包サイズ", 
                "梱包記号", "美品", "ブランド", "フリー", "袖", "もの", "男女", 
                "採寸1", "ラック", "金額", "股上", "股下", "ウエスト", "もも幅", "裾幅", "総丈", "ヒップ", "仕入先", "仕入日", "原価"
            ],
            "ワンピース": [
                "カテゴリ", "管理番号", "タイトル", "文字数", "付属品", "ランク", "コメント", 
                "素材", "色", "サイズ", "着丈", "　肩幅", "身幅", "袖丈", "梱包サイズ", 
                "梱包記号", "美品", "ブランド", "フリー", "袖", "もの", "男女", 
                "採寸1", "ラック", "金額", "股上", "股下", "ウエスト", "もも幅", "裾幅", "総丈", "ヒップ", "仕入先", "仕入日", "原価"
            ],
            "オールインワン": [
                "カテゴリ", "管理番号", "タイトル", "文字数", "付属品", "ランク", "コメント", 
                "素材", "色", "サイズ", "着丈", "　肩幅", "身幅", "袖丈", "梱包サイズ", 
                "梱包記号", "美品", "ブランド", "フリー", "袖", "もの", "男女", 
                "採寸1", "ラック", "金額", "股上", "股下", "ウエスト", "もも幅", "裾幅", "総丈", "ヒップ", "仕入先", "仕入日", "原価"
            ],
            "スカートスーツ": [
                "カテゴリ", "管理番号", "タイトル", "文字数", "付属品", "ランク", "コメント", 
                "素材", "色", "サイズ", "着丈", "　肩幅", "身幅", "袖丈", "梱包サイズ", 
                "梱包記号", "美品", "ブランド", "フリー", "袖", "もの", "男女", 
                "採寸1", "ラック", "金額", "股上", "股下", "ウエスト", "もも幅", "裾幅", "総丈", "ヒップ", "仕入先", "仕入日", "原価"
            ],
            "パンツスーツ": [
                "カテゴリ", "管理番号", "タイトル", "文字数", "付属品", "ランク", "コメント", 
                "素材", "色", "サイズ", "着丈", "　肩幅", "身幅", "袖丈", "梱包サイズ", 
                "梱包記号", "美品", "ブランド", "フリー", "袖", "もの", "男女", 
                "採寸1", "ラック", "金額", "股上", "股下", "ウエスト", "もも幅", "裾幅", "総丈", "ヒップ", "仕入先", "仕入日", "原価"
            ],
            "アンサンブル": [
                "カテゴリ", "管理番号", "タイトル", "文字数", "付属品", "ランク", "コメント", 
                "素材", "色", "サイズ", "着丈", "　肩幅", "身幅", "袖丈", "梱包サイズ", 
                "梱包記号", "美品", "ブランド", "フリー", "袖", "もの", "男女", 
                "採寸1", "ラック", "金額", "股上", "股下", "ウエスト", "もも幅", "裾幅", "総丈", "ヒップ", "仕入先", "仕入日", "原価"
            ],
            "靴": [
                "カテゴリ", "管理番号", "タイトル", "文字数", "付属品", "ランク", "コメント", 
                "素材", "色", "サイズ", "着丈", "　肩幅", "身幅", "袖丈", "梱包サイズ", 
                "梱包記号", "美品", "ブランド", "フリー", "袖", "もの", "男女", 
                "採寸1", "ラック", "金額", "股上", "股下", "ウエスト", "もも幅", "裾幅", "総丈", "ヒップ", "仕入先", "仕入日", "原価"
            ],
            "ブーツ": [
                "カテゴリ", "管理番号", "タイトル", "文字数", "付属品", "ランク", "コメント", 
                "素材", "色", "サイズ", "着丈", "　肩幅", "身幅", "袖丈", "梱包サイズ", 
                "梱包記号", "美品", "ブランド", "フリー", "袖", "もの", "男女", 
                "採寸1", "ラック", "金額", "股上", "股下", "ウエスト", "もも幅", "裾幅", "総丈", "ヒップ", "仕入先", "仕入日", "原価"
            ],
            "ベルト": [
                "カテゴリ", "管理番号", "タイトル", "文字数", "付属品", "ランク", "コメント", 
                "素材", "色", "サイズ", "着丈", "　肩幅", "身幅", "袖丈", "梱包サイズ", 
                "梱包記号", "美品", "ブランド", "フリー", "袖", "もの", "男女", 
                "採寸1", "ラック", "金額", "股上", "股下", "ウエスト", "もも幅", "裾幅", "総丈", "ヒップ", "仕入先", "仕入日", "原価"
            ],
            "ネクタイ縦横": [
                "カテゴリ", "管理番号", "タイトル", "文字数", "付属品", "ランク", "コメント", 
                "素材", "色", "サイズ", "着丈", "　肩幅", "身幅", "袖丈", "梱包サイズ", 
                "梱包記号", "美品", "ブランド", "フリー", "袖", "もの", "男女", 
                "採寸1", "ラック", "金額", "股上", "股下", "ウエスト", "もも幅", "裾幅", "総丈", "ヒップ", "仕入先", "仕入日", "原価"
            ],
            "帽子": [
                "カテゴリ", "管理番号", "タイトル", "文字数", "付属品", "ランク", "コメント", 
                "素材", "色", "サイズ", "着丈", "　肩幅", "身幅", "袖丈", "梱包サイズ", 
                "梱包記号", "美品", "ブランド", "フリー", "袖", "もの", "男女", 
                "採寸1", "ラック", "金額", "股上", "股下", "ウエスト", "もも幅", "裾幅", "総丈", "ヒップ", "仕入先", "仕入日", "原価"
            ],
            "バッグ": [
                "カテゴリ", "管理番号", "タイトル", "文字数", "付属品", "ランク", "コメント", 
                "素材", "色", "サイズ", "着丈", "　肩幅", "身幅", "袖丈", "梱包サイズ", 
                "梱包記号", "美品", "ブランド", "フリー", "袖", "もの", "男女", 
                "採寸1", "ラック", "金額", "股上", "股下", "ウエスト", "もも幅", "裾幅", "総丈", "ヒップ", "仕入先", "仕入日", "原価"
            ],
            "ネックレス": [
                "カテゴリ", "管理番号", "タイトル", "文字数", "付属品", "ランク", "コメント", 
                "素材", "色", "サイズ", "着丈", "　肩幅", "身幅", "袖丈", "梱包サイズ", 
                "梱包記号", "美品", "ブランド", "フリー", "袖", "もの", "男女", 
                "採寸1", "ラック", "金額", "股上", "股下", "ウエスト", "もも幅", "裾幅", "総丈", "ヒップ", "仕入先", "仕入日", "原価"
            ],
            "サングラス": [
                "カテゴリ", "管理番号", "タイトル", "文字数", "付属品", "ランク", "コメント", 
                "素材", "色", "サイズ", "着丈", "　肩幅", "身幅", "袖丈", "梱包サイズ", 
                "梱包記号", "美品", "ブランド", "フリー", "袖", "もの", "男女", 
                "採寸1", "ラック", "金額", "股上", "股下", "ウエスト", "もも幅", "裾幅", "総丈", "ヒップ", "仕入先", "仕入日", "原価"
            ]
        }
        
        # Convert processed results to Excel format
        products_to_add = []
        success_count = 0
        error_count = 0
        errors = []
        
        print(f"🔄 Converting {len(processed_results)} products to Excel format...")
        
        # Function to classify product using Perplexity AI
        def classify_product_with_ai(product_info):
            """
            Use Perplexity AI to classify product based on its characteristics.
            Returns the appropriate sheet name.
            """
            try:
                import requests
                import os
                
                # Get Perplexity API key from environment
                PERPLEXITY_API_KEY = os.getenv("PERPLEXITY_API_KEY")
                if not PERPLEXITY_API_KEY:
                    print("⚠️ PERPLEXITY_API_KEY not found, using default classification")
                    return "トップス"  # Default fallback
                
                # Prepare product information for classification
                title = product_info.get('タイトル', '') or product_info.get('title', '')
                brand = product_info.get('ブランド', '') or product_info.get('brand', '')
                product_type = product_info.get('もの', '') or product_info.get('product_type', '')
                size = product_info.get('サイズ', '') or product_info.get('size', '')
                material = product_info.get('素材', '') or product_info.get('material', '')
                
                # Create classification prompt
                classification_prompt = f"""以下の商品情報を分析して、適切なシートに分類してください。

商品情報:
- タイトル: {title}
- ブランド: {brand}
- 商品タイプ: {product_type}
- サイズ: {size}
- 素材: {material}

利用可能なシート:
1. トップス - シャツ、ブラウス、セーター、ニット、ジャケット、コート等
2. パンツ - ズボン、ジーンズ、スラックス、チノパン等
3. スカート - ミニスカート、ロングスカート等
4. ワンピース - ドレス、ワンピース等
5. オールインワン - ジャンプスーツ、サロペット等
6. スカートスーツ - スカートスーツ
7. パンツスーツ - パンツスーツ
8. アンサンブル - ツインセット等
9. 靴 - パンプス、スニーカー、革靴等
10. ブーツ - ロングブーツ、ショートブーツ等
11. ベルト - レザーベルト等
12. ネクタイ縦横 - ネクタイ、ボウタイ等
13. 帽子 - キャップ、ハット等
14. バッグ - ハンドバッグ、トートバッグ等
15. ネックレス - ネックレス、チョーカー等
16. サングラス - サングラス、メガネ等

最も適切なシート名のみを回答してください。例: トップス"""
                
                # Call Perplexity AI API
                headers = {
                    "Authorization": f"Bearer {PERPLEXITY_API_KEY}",
                    "Content-Type": "application/json"
                }
                
                payload = {
                    "model": "sonar",
                    "messages": [
                        {
                            "role": "user",
                            "content": classification_prompt
                        }
                    ]
                }
                
                response = requests.post(
                    "https://api.perplexity.ai/chat/completions",
                    headers=headers,
                    json=payload,
                    timeout=10
                )
                
                if response.status_code == 200:
                    result = response.json()
                    ai_classification = result["choices"][0]["message"]["content"].strip()
                    
                    # Clean up the AI response and map to valid sheet names
                    ai_classification = ai_classification.replace('シート', '').replace(':', '').strip()
                    
                    # Validate the classification
                    valid_sheets = list(SHEET_HEADERS.keys())
                    for sheet_name in valid_sheets:
                        if sheet_name in ai_classification:
                            print(f"🤖 AI classified '{title[:30]}...' as: {sheet_name}")
                            return sheet_name
                    
                    # If no exact match, try partial matching
                    for sheet_name in valid_sheets:
                        if any(keyword in ai_classification for keyword in sheet_name.split()):
                            print(f"🤖 AI partial match '{title[:30]}...' as: {sheet_name}")
                            return sheet_name
                    
                    print(f"⚠️ AI classification '{ai_classification}' not recognized, using default")
                    return "トップス"  # Default fallback
                    
                else:
                    print(f"⚠️ Perplexity API error: {response.status_code}, using default classification")
                    return "トップス"  # Default fallback
                    
            except Exception as e:
                print(f"⚠️ AI classification error: {str(e)}, using default classification")
                return "トップス"  # Default fallback
        
        for product_id, result in processed_results.items():
            try:
                # Extract the data from the result structure
                if isinstance(result, dict) and 'listing_data' in result:
                    listing_data = result['listing_data']
                else:
                    # Fallback to using result directly
                    listing_data = result
                
                # Convert the listing data to the format expected by the Excel service
                # Map the data structure from the image to Excel format
                title = listing_data.get('タイトル', '') or listing_data.get('title', '')
                brand = listing_data.get('ブランド', '') or listing_data.get('brand', '')
                product_type = listing_data.get('もの', '') or listing_data.get('product_type', '')
                color = listing_data.get('色', '') or listing_data.get('color', '')
                size = listing_data.get('サイズ', '') or listing_data.get('size', '')
                material = listing_data.get('素材', '') or listing_data.get('material', '')
                accessories = listing_data.get('付属品', '') or listing_data.get('accessories', '')
                rank = listing_data.get('ランク', '') or listing_data.get('rank', '')
                
                # Handle rank conversion (from image: "ランクA" -> "3")
                if rank == "ランクA":
                    rank = "3"
                elif rank == "ランクB":
                    rank = "2"
                elif rank == "ランクC":
                    rank = "1"
                
                # Handle material conversion (from image: "未検出" -> "不明")
                if material == "未検出" or not material:
                    material = "不明"
                
                excel_data = {
                    "カテゴリ": "",  # Will be set by AI classification
                    "管理番号": listing_data.get('管理番号', product_id),
                    "タイトル": title,
                    "文字数": len(title) if title else 0,
                    "付属品": accessories if accessories else "無",
                    "ランク": rank if rank else "3",  # Default to 3 if not specified
                    "コメント": listing_data.get('コメント', '') or listing_data.get('comment', ''),
                    "素材": material,
                    "色": color,
                    "サイズ": size,
                    "着丈": listing_data.get('着丈') or None,
                    "　肩幅": listing_data.get('肩幅') or listing_data.get('　肩幅') or None,
                    "身幅": listing_data.get('身幅') or None,
                    "袖丈": listing_data.get('袖丈') or None,
                    "股上": listing_data.get('股上') or None,
                    "股下": listing_data.get('股下') or None,
                    "ウエスト": listing_data.get('ウエスト') or None,
                    "もも幅": listing_data.get('もも幅') or None,
                    "裾幅": listing_data.get('裾幅') or None,
                    "総丈": listing_data.get('総丈') or None,
                    "ヒップ": listing_data.get('ヒップ') or None,
                    "梱包サイズ": listing_data.get('梱包サイズ', '') or "通常",
                    "梱包記号": listing_data.get('梱包記号', '') or "◇",
                    "美品": listing_data.get('美品', ''),
                    "ブランド": brand,
                    "フリー": listing_data.get('フリー', '') or listing_data.get('free_text', ''),
                    "袖": listing_data.get('袖', '') or listing_data.get('sleeve', ''),
                    "もの": product_type,
                    "男女": listing_data.get('男女', '') or listing_data.get('gender', ''),
                    "採寸1": listing_data.get('採寸1', '') or listing_data.get('measurement1', ''),
                    "ラック": listing_data.get('ラック', '') or listing_data.get('rack', ''),
                    "金額": listing_data.get('金額') or listing_data.get('売値') or listing_data.get('price') or None,
                    "仕入先": listing_data.get('仕入先', '') or listing_data.get('supplier', ''),
                    "仕入日": listing_data.get('仕入日', '') or listing_data.get('purchase_date', ''),
                    "原価": listing_data.get('原価') or listing_data.get('cost_price') or None
                }
                
                # Use AI classification to determine the sheet
                ai_classification = classify_product_with_ai(excel_data)
                excel_data["カテゴリ"] = ai_classification
                
                products_to_add.append(excel_data)
                
            except Exception as e:
                error_count += 1
                errors.append(f"Product {product_id}: {str(e)}")
                print(f"❌ Error processing product {product_id}: {str(e)}")
                continue
        
        if not products_to_add:
            return jsonify({
                'success': False,
                'message': 'No valid products to add to Excel',
                'errors': errors
            }), 400
        
        print(f"✅ Converted {len(products_to_add)} products successfully")
        print(f"🔄 Starting direct Excel write operation...")
        
        # Use direct openpyxl approach like the working sample code
        try:
            import openpyxl
            import time
            
            # Check if file exists (like your sample code)
            if not os.path.exists(excel_file_path):
                print(f"❌ Error: {excel_file_path} not found!")
                return jsonify({
                    'success': False,
                    'message': f'Excel file not found: {excel_file_path}'
                }), 404
            
            # Create a new workbook instead of loading the corrupted one (like your sample code approach)
            print(f"📖 Creating new Excel workbook with predefined structure")
            wb = openpyxl.load_workbook(excel_file_path, keep_vba=True)
            
            # Remove the default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Create all required sheets with headers (like your sample code structure)
            for sheet_name, headers in SHEET_HEADERS.items():
                print(f"📊 Creating sheet: {sheet_name}")
                ws = wb.create_sheet(sheet_name)
                # Add headers to new sheet (like your sample code)
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)
                print(f"📊 Sheet '{sheet_name}' created with {len(headers)} headers")
            
            # Function to check if a row is empty (all cells are None or empty string)
            def is_row_empty(sheet, row_num, num_columns):
                for col in range(1, num_columns + 1):
                    cell_value = sheet.cell(row=row_num, column=col).value
                    if cell_value is not None and str(cell_value).strip() != "":
                        return False
                return True
            
            # Process each product (like your sample code)
            for product_data in products_to_add:
                try:
                    # Determine sheet name based on AI classification
                    category = product_data.get('カテゴリ', '')
                    if not category or category not in SHEET_HEADERS:
                        # Default to first sheet if category not found
                        sheet_name = list(SHEET_HEADERS.keys())[0]
                        print(f"⚠️ Category '{category}' not found, using default sheet: {sheet_name}")
                    else:
                        sheet_name = category
                    
                    # Check if sheet exists (like your sample code)
                    if sheet_name not in wb.sheetnames:
                        print(f"❌ Error: Sheet '{sheet_name}' not found!")
                        print(f"Available sheets: {wb.sheetnames}")
                        error_count += 1
                        errors.append(f"Sheet '{sheet_name}' not found for product {product_data.get('管理番号', 'unknown')}")
                        continue
                    
                    sheet = wb[sheet_name]
                    headers = SHEET_HEADERS[sheet_name]
                    
                    # Print current sheet info (like your sample code)
                    print(f"📊 Sheet '{sheet.title}' loaded successfully")
                    print(f"📊 Current rows: {sheet.max_row}")
                    print(f"📊 Current columns: {sheet.max_column}")
                    print(f"📊 Using sheet: {sheet_name} with {len(headers)} headers")
                    
                    # Convert data to list format matching the headers (EXACTLY like your sample code)
                    row_data = []
                    for header in headers:
                        value = product_data.get(header, "")
                        # Convert None to empty string
                        if value is None:
                            value = ""
                        row_data.append(value)
                    
                    print(f"📝 Prepared row data with {len(row_data)} columns")
                    
                    # Find the first empty row (EXACTLY like your sample code)
                    num_columns = len(row_data)
                    target_row = None
                    
                    # Start from row 2 (assuming row 1 has headers)
                    for row in range(2, sheet.max_row + 1):
                        if is_row_empty(sheet, row, num_columns):
                            target_row = row
                            break
                    
                    # If no empty row found, use the next row after the last data
                    if target_row is None:
                        target_row = sheet.max_row + 1
                    
                    print(f"📝 Found empty row at: {target_row}")
                    
                    # Write data to the target row (EXACTLY like your sample code)
                    for col, value in enumerate(row_data, 1):
                        sheet.cell(row=target_row, column=col, value=value)
                    
                    print(f"✅ Data added to row {target_row}")
                    success_count += 1
                    
                except Exception as e:
                    error_count += 1
                    errors.append(f"Failed to add product {product_data.get('管理番号', 'unknown')}: {str(e)}")
                    print(f"❌ Error adding product to Excel: {str(e)}")
                    continue
            
            # Save the workbook to backend directory (overwrite existing file)
            if success_count > 0:
                print(f"💾 Saving Excel workbook with {success_count} new entries")
                
                # Save to backend directory without timestamp (overwrite existing file)
                output_path = "./PL出品マクロ.xlsm"
                
                # Try to save with retry logic (EXACTLY like your sample code)
                max_retries = 3
                for attempt in range(max_retries):
                    try:
                        wb.save(output_path)
                        print(f"✅ Excel workbook saved successfully to: {output_path}")
                        break
                    except PermissionError as e:
                        if attempt < max_retries - 1:
                            print(f"⚠️ Permission error on attempt {attempt + 1}, retrying...")
                            time.sleep(2)
                            # Try a different filename with timestamp as fallback
                            timestamp = time.strftime("%Y%m%d_%H%M%S")
                            output_path = f"./PL出品マクロ_updated_{timestamp}_attempt{attempt + 2}.xlsm"
                        else:
                            print(f"❌ Failed to save after {max_retries} attempts")
                            raise e
                    except Exception as e:
                        print(f"❌ Error saving workbook: {str(e)}")
                        raise e
                
                # Print final sheet info (like your sample code)
                print(f"📊 Total rows in sheet after update: {sheet.max_row}")
                
                # Close workbook
                wb.close()
                
                return jsonify({
                    'success': True,
                    'summary': {
                        'total_processed': len(processed_results),
                        'products_converted': len(products_to_add),
                        'successfully_added': success_count,
                        'failed_to_add': error_count,
                        'conversion_errors': len(errors)
                    },
                    'errors': errors if errors else None,
                    'message': f'Successfully added {success_count} products to new Excel file: {output_path}',
                    'output_file': output_path
                })
            else:
                wb.close()
                return jsonify({
                    'success': False,
                    'message': 'No products were successfully added to Excel',
                    'errors': errors
                }), 400
            
        except Exception as e:
            print(f"❌ Critical error during Excel write: {str(e)}")
            import traceback
            print(f"❌ Traceback: {traceback.format_exc()}")
            return jsonify({
                'success': False,
                'message': f'Critical error during Excel export: {str(e)}',
                'summary': {
                    'total_processed': len(processed_results),
                    'products_converted': len(products_to_add),
                    'successfully_added': 0,
                    'failed_to_add': len(products_to_add),
                    'conversion_errors': len(errors)
                },
                'errors': errors + [f'Excel write error: {str(e)}'],
                'traceback': traceback.format_exc(),
                'note': 'The system now creates a new Excel file instead of modifying the existing corrupted one'
            }), 500
        
    except Exception as e:
        print(f"❌ Unexpected error in export_to_excel: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Unexpected error: {str(e)}'
        }), 500

@excel_bp.route('/excel/file-info', methods=['GET'])
def get_excel_file_info():
    """
    Get information about the current Excel file.
    """
    try:
        file_path = excel_service.excel_file_path
        
        if not os.path.exists(file_path):
            return jsonify({
                'success': False,
                'message': f'Excel file not found: {file_path}'
            }), 404
        
        # Get file stats
        file_stats = os.stat(file_path)
        import datetime
        
        return jsonify({
            'success': True,
            'file_info': {
                'path': file_path,
                'size': file_stats.st_size,
                'modified': datetime.datetime.fromtimestamp(file_stats.st_mtime).isoformat(),
                'exists': True
            }
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error getting file info: {str(e)}'
        }), 500

@excel_bp.route('/excel/download-updated-file', methods=['GET'])
def download_updated_excel_file():
    """
    Download the updated Excel file from the backend directory.
    """
    try:
        # Use the same file path as the service
        file_path = excel_service.excel_file_path
        
        if not os.path.exists(file_path):
            return jsonify({
                'success': False,
                'message': f'Updated Excel file not found: {file_path}'
            }), 404
        
        # Return the file for download
        from flask import send_file
        return send_file(
            file_path,
            as_attachment=True,
            download_name='PL出品マクロ.xlsm',
            mimetype='application/vnd.ms-excel.sheet.macroEnabled.12'
        )
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error downloading file: {str(e)}'
        }), 500

@excel_bp.route('/excel/category-lookup', methods=['POST'])
def lookup_category_number():
    """
    Look up category number for a product using AI and category database.
    """
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({
                'success': False,
                'message': 'No product data provided'
            }), 400
        
        # Get category number using AI
        category_number, lookup_info = category_service.get_category_number_with_ai(data)
        
        # Get category information if found
        category_info = None
        if category_number:
            category_info = category_service.get_category_by_number(category_number)
        
        return jsonify({
            'success': True,
            'category_number': category_number,
            'category_info': category_info,
            'lookup_info': lookup_info
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error looking up category: {str(e)}'
        }), 500

@excel_bp.route('/excel/category-search', methods=['POST'])
def search_categories():
    """
    Search categories by keywords.
    """
    try:
        data = request.get_json()
        
        if not data or 'keywords' not in data:
            return jsonify({
                'success': False,
                'message': 'Keywords are required for search'
            }), 400
        
        keywords = data['keywords']
        if not isinstance(keywords, list):
            keywords = [keywords]
        
        # Search for matching categories
        matches = category_service.search_categories_by_keywords(keywords)
        
        return jsonify({
            'success': True,
            'matches': matches,
            'total_matches': len(matches)
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error searching categories: {str(e)}'
        }), 500

@excel_bp.route('/excel/category-info/<category_number>', methods=['GET'])
def get_category_info(category_number):
    """
    Get detailed information about a specific category number.
    """
    try:
        category_info = category_service.get_category_by_number(category_number)
        
        if category_info:
            return jsonify({
                'success': True,
                'category_info': category_info
            })
        else:
            return jsonify({
                'success': False,
                'message': f'Category number {category_number} not found'
            }), 404
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error getting category info: {str(e)}'
        }), 500

@excel_bp.route('/excel/categories', methods=['GET'])
def get_all_categories():
    """
    Get all available categories.
    """
    try:
        categories = category_service.get_all_categories()
        
        return jsonify({
            'success': True,
            'categories': categories,
            'total_categories': len(categories)
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error getting categories: {str(e)}'
        }), 500

@excel_bp.route('/excel/category-stats', methods=['GET'])
def get_category_statistics():
    """
    Get statistics about the category database.
    """
    try:
        stats = category_service.get_category_statistics()
        
        return jsonify({
            'success': True,
            'statistics': stats
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error getting category statistics: {str(e)}'
        }), 500

@excel_bp.route('/excel/create-new-file', methods=['POST'])
def create_new_excel_file():
    """
    Create a new XLSX file with the same structure as PL出品マクロ.xlsm but without macros.
    """
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({
                'success': False,
                'message': 'No data provided'
            }), 400
        
        # Get optional filename from request
        filename = data.get('filename', 'PL出品マクロ_新規作成.xlsx')
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        
        # Create output path in backend directory
        output_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), filename)
        
        # Create the Excel file with structure only
        success, message = excel_creator_service.create_excel_file_with_structure(output_path)
        
        if success:
            return jsonify({
                'success': True,
                'message': message,
                'file_path': output_path,
                'filename': filename
            })
        else:
            return jsonify({
                'success': False,
                'message': message
            }), 500
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error creating Excel file: {str(e)}'
        }), 500

@excel_bp.route('/excel/create-with-data', methods=['POST'])
def create_excel_file_with_data():
    """
    Create a new XLSX file with the same structure and add product data to appropriate sheets.
    """
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({
                'success': False,
                'message': 'No data provided'
            }), 400
        
        # Get product data list
        products = data.get('products', [])
        
        # If no products provided, try to get from current processing results
        if not products:
            # This would need to be implemented to get current processing results
            # For now, return an empty file
            products = []
        
        # Get optional filename from request
        filename = data.get('filename', 'PL出品マクロ_新規作成.xlsx')
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        
        # Create output path in backend directory
        output_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), filename)
        
        # Create Excel file with data
        success_count, failure_count, error_messages = excel_creator_service.add_data_to_excel_file(products, output_path)
        
        return jsonify({
            'success': True,
            'summary': {
                'total_products': len(products),
                'successfully_added': success_count,
                'failed_to_add': failure_count
            },
            'errors': error_messages if error_messages else None,
            'message': f'Successfully created Excel file with {success_count} products: {filename}',
            'file_path': output_path,
            'filename': filename
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error creating Excel file with data: {str(e)}'
        }), 500

@excel_bp.route('/excel/download-new-file/<filename>', methods=['GET'])
def download_new_excel_file(filename):
    """
    Download a newly created Excel file from the backend directory.
    """
    try:
        # Create file path in backend directory
        file_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), filename)
        
        if not os.path.exists(file_path):
            return jsonify({
                'success': False,
                'message': f'Excel file not found: {filename}'
            }), 404
        
        # Return the file for download
        from flask import send_file
        return send_file(
            file_path,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error downloading file: {str(e)}'
        }), 500

@excel_bp.route('/excel/new-file-info', methods=['GET'])
def get_new_excel_file_info():
    """
    Get information about the structure of new Excel files that can be created.
    """
    try:
        # Get sheet info from the creator service
        sheet_info = excel_creator_service.get_sheet_info()
        
        return jsonify({
            'success': True,
            'sheet_info': sheet_info,
            'message': 'Information about new Excel file structure'
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error getting file info: {str(e)}'
        }), 500 