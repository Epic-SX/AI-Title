"""
Excel creator service for creating new XLSX files with the same structure as PL出品マクロ.xlsm
but without macros, allowing for clean data export.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import Dict, List, Tuple, Optional
import os
import time
import logging
from app.services.category_lookup_service import CategoryLookupService
from app.utils.excel_utils import get_predefined_sheet_headers

# Set up logging
logger = logging.getLogger(__name__)

class ExcelCreatorService:
    def __init__(self):
        # Initialize category lookup service
        self.category_service = CategoryLookupService()
        
        # Get predefined sheet headers from utils
        self.SHEET_HEADERS = get_predefined_sheet_headers()
        
        # Category classification keywords for fallback
        self.category_keywords = {
            'トップス': [
                'ブラウス', 'シャツ', 'tシャツ', 'カットソー', 'ニット', 'セーター', 
                'パーカー', 'フリース', 'ジャケット', 'カーディガン', 'ベスト',
                'タンクトップ', 'キャミソール', 'チュニック'
            ],
            'パンツ': [
                'パンツ', 'ズボン', 'ジーンズ', 'デニム', 'チノパン', 'スラックス',
                'レギンス', 'ショートパンツ', 'ハーフパンツ', 'ワイドパンツ',
                'スキニー', 'ボトムス', 'トラウザー'
            ],
            'スカート': [
                'スカート', 'ミニスカート', 'ロングスカート', 'マキシスカート',
                'フレアスカート', 'タイトスカート', 'プリーツスカート'
            ],
            'ワンピース': [
                'ワンピース', 'ドレス', 'マキシワンピース', 'ミニワンピース',
                'シャツワンピース', 'ニットワンピース'
            ],
            'オールインワン': [
                'オールインワン', 'サロペット', 'オーバーオール', 'ジャンプスーツ',
                'コンビネゾン', 'つなぎ'
            ],
            'スカートスーツ': [
                'スカートスーツ', 'スーツ.*スカート', 'セットアップ.*スカート'
            ],
            'パンツスーツ': [
                'パンツスーツ', 'スーツ.*パンツ', 'セットアップ.*パンツ'
            ],
            'アンサンブル': [
                'アンサンブル', 'ツインセット', 'セット.*ニット'
            ],
            '靴': [
                'パンプス', 'ヒール', 'フラットシューズ', '革靴', 'ローファー',
                'サンダル', 'ミュール', 'オックスフォード'
            ],
            'ブーツ': [
                'ブーツ', 'ロングブーツ', 'ショートブーツ', 'アンクルブーツ',
                'ニーハイブーツ', 'ムートンブーツ'
            ],
            'ベルト': [
                'ベルト', 'レザーベルト', 'チェーンベルト'
            ],
            'ネクタイ縦横': [
                'ネクタイ', 'タイ', 'ボウタイ'
            ],
            '帽子': [
                '帽子', 'キャップ', 'ハット', 'ベレー帽', 'ニット帽',
                'ビーニー', '麦わら帽子', 'ハンチング'
            ],
            'バッグ': [
                'バッグ', 'ハンドバッグ', 'ショルダーバッグ', 'トートバッグ',
                'クラッチバッグ', 'リュック', 'バックパック', 'ポーチ',
                'ウエストバッグ', 'メッセンジャーバッグ'
            ],
            'ネックレス': [
                'ネックレス', 'チョーカー', 'ペンダント', 'チェーン'
            ],
            'サングラス': [
                'サングラス', 'メガネ', '眼鏡', 'グラス'
            ]
        }
        
        # Default sheet for unclassified items
        self.default_sheet = 'トップス'
    
    def get_category_number_for_product(self, product_data: Dict) -> Tuple[Optional[str], Dict]:
        """
        Get category number for a product using AI and category lookup service.
        
        Args:
            product_data: Dictionary containing product information
            
        Returns:
            Tuple of (category_number, lookup_info)
        """
        try:
            # Use the category lookup service to get category number
            category_number, lookup_info = self.category_service.get_category_number_with_ai(product_data)
            
            logger.info(f"[CATEGORY] Lookup for product: {product_data.get('title', 'Unknown')} -> {category_number}")
            
            return category_number, lookup_info
            
        except Exception as e:
            logger.error(f"❌ Error getting category number: {str(e)}")
            return None, {'method': 'error', 'error': str(e)}
    
    def classify_product_category(self, title: str, product_data: Dict) -> str:
        """
        Classify product category based on title and product data.
        Returns the appropriate sheet name.
        """
        import re
        title_lower = title.lower()
        
        # Check if product_data has specific type information
        product_type = product_data.get('もの', '') or product_data.get('product_type', '')
        if product_type:
            title_lower += ' ' + product_type.lower()
        
        # Check for keywords in each category
        for category, keywords in self.category_keywords.items():
            for keyword in keywords:
                if re.search(keyword.lower(), title_lower):
                    return category
        
        # If no specific category found, return default
        return self.default_sheet
    
    def map_data_to_sheet_headers(self, data: Dict, sheet_name: str) -> Dict:
        """
        Map the input data to the appropriate headers for the target sheet.
        Uses predefined headers instead of reading from Excel file.
        """
        try:
            # Use predefined headers instead of reading from Excel file
            if sheet_name not in self.SHEET_HEADERS:
                logger.error(f"Sheet '{sheet_name}' not found in predefined headers")
                return {}
            
            headers = self.SHEET_HEADERS[sheet_name]
            
        except Exception as e:
            logger.error(f"Error getting sheet headers: {e}")
            return {}
        
        # Create mapping based on common field names
        field_mappings = {
            'カテゴリ': ['category', 'カテゴリ'],
            '管理番号': ['management_number', '管理番号', 'id'],
            'タイトル': ['title', 'タイトル'],
            '文字数': ['character_count', '文字数'],
            '付属品': ['accessories', '付属品'],
            '日本サイズ': ['japanese_size', '日本サイズ'],
            'ランク': ['rank', 'ランク', 'condition_rank'],
            'コメント': ['comment', 'コメント', 'description'],
            '仕立て・収納': ['tailoring_storage', '仕立て・収納'],
            '素材': ['material', '素材'],
            '色': ['color', '色'],
            'サイズ': ['size', 'サイズ'],
            '梱包サイズ': ['packaging_size', '梱包サイズ'],
            '梱包記号': ['packaging_symbol', '梱包記号'],
            '美品': ['excellent_condition', '美品'],
            'ブランド': ['brand', 'ブランド'],
            'フリー': ['free_text', 'フリー'],
            '袖': ['sleeve', '袖'],
            'もの': ['item_type', 'もの', 'product_type'],
            '男女': ['gender', '男女'],
            'ラック': ['rack', 'ラック'],
            '仕入先': ['supplier', '仕入先'],
            '仕入日': ['purchase_date', '仕入日'],
            '原価': ['cost_price', '原価'],
            '金額': ['price', '金額', 'amount'],
            # Measurement fields
            '着丈': ['garment_length', '着丈'],
            '　肩幅': ['shoulder_width', 'shoulder_width', '肩幅', '　肩幅'],
            '身幅': ['chest_width', '身幅'],
            '袖丈': ['sleeve_length', '袖丈'],
            '股上': ['rise', '股上'],
            '股下': ['inseam', '股下'],
            'ウエスト': ['waist', 'ウエスト'],
            'もも幅': ['thigh_width', 'もも幅'],
            '裾幅': ['hem_width', '裾幅'],
            '総丈': ['total_length', '総丈'],
            'ヒップ': ['hip', 'ヒップ'],
            '採寸1': ['measurement1', '採寸1'],
            '採寸2': ['measurement2', '採寸2']
        }
        
        # Create the mapped data
        mapped_data = {}
        
        for header in headers:
            # Find matching field in input data
            value = None
            
            # Direct match
            if header in data:
                value = data[header]
            else:
                # Check mappings
                if header in field_mappings:
                    for field_key in field_mappings[header]:
                        if field_key in data:
                            value = data[field_key]
                            break
            
            # Set the value (None will be handled as empty)
            mapped_data[header] = value
        
        return mapped_data
    
    def generate_measurement_text(self, data: Dict, sheet_name: str) -> str:
        """
        Generate formatted measurement text based on the product category.
        """
        measurements = []
        
        if sheet_name == 'トップス':
            if data.get('着丈'):
                measurements.append(f"着丈：約{data['着丈']}cm")
            if data.get('　肩幅') or data.get('肩幅'):
                shoulder = data.get('　肩幅') or data.get('肩幅')
                measurements.append(f"肩幅：約{shoulder}cm")
            if data.get('身幅'):
                measurements.append(f"身幅：約{data['身幅']}cm")
            if data.get('袖丈'):
                measurements.append(f"袖丈：約{data['袖丈']}cm")
                
        elif sheet_name == 'パンツ':
            if data.get('股上'):
                measurements.append(f"股上：約{data['股上']}cm")
            if data.get('股下'):
                measurements.append(f"股下：約{data['股下']}cm")
            if data.get('ウエスト'):
                measurements.append(f"ウエスト：約{data['ウエスト']}cm")
            if data.get('もも幅'):
                measurements.append(f"もも幅：約{data['もも幅']}cm")
            if data.get('裾幅'):
                measurements.append(f"裾幅：約{data['裾幅']}cm")
                
        elif sheet_name == 'スカート':
            if data.get('総丈'):
                measurements.append(f"総丈：約{data['総丈']}cm")
            if data.get('ウエスト'):
                measurements.append(f"ウエスト：約{data['ウエスト']}cm")
            if data.get('ヒップ'):
                measurements.append(f"ヒップ：約{data['ヒップ']}cm")
        
        return "　".join(measurements) if measurements else ""
    
    def create_excel_file_with_structure(self, output_path: str) -> Tuple[bool, str]:
        """
        Create a new Excel file with the same structure as PL出品マクロ.xlsm but without macros.
        
        Args:
            output_path: Path where the new Excel file should be created
            
        Returns:
            Tuple of (success: bool, message: str)
        """
        try:
            # Create a new workbook
            wb = Workbook()
            
            # Remove the default sheet
            wb.remove(wb.active)
            
            # Create all sheets with headers
            for sheet_name, headers in self.SHEET_HEADERS.items():
                ws = wb.create_sheet(title=sheet_name)
                
                # Add headers to the first row
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    
                    # Style the header row
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    # Add border
                    thin_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    cell.border = thin_border
                
                # Set column widths (adjust as needed)
                for col in range(1, len(headers) + 1):
                    ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 15
            
            # Save the workbook
            wb.save(output_path)
            wb.close()
            
            logger.info(f"✅ Created new Excel file with structure: {output_path}")
            return True, f"Successfully created Excel file: {output_path}"
            
        except Exception as e:
            logger.error(f"❌ Error creating Excel file: {str(e)}")
            return False, f"Error creating Excel file: {str(e)}"
    
    def add_data_to_excel_file(self, data_list: List[Dict], output_path: str) -> Tuple[int, int, List[str]]:
        """
        Create a new Excel file with the same structure and add product data to appropriate sheets.
        
        Args:
            data_list: List of dictionaries containing product data
            output_path: Path where the new Excel file should be created
            
        Returns:
            Tuple of (success_count: int, failure_count: int, error_messages: List[str])
        """
        if not data_list:
            return 0, 0, []
        
        success_count = 0
        failure_count = 0
        error_messages = []
        
        try:
            # First create the Excel file with structure
            success, message = self.create_excel_file_with_structure(output_path)
            if not success:
                return 0, len(data_list), [f"Failed to create Excel file: {message}"]
            
            # Now load the created file and add data
            from openpyxl import load_workbook
            wb = load_workbook(output_path)
            
            # Process all products
            for i, data in enumerate(data_list):
                try:
                    # Classify the product to determine target sheet
                    title = data.get('タイトル', '') or data.get('title', '')
                    if not title:
                        error_messages.append(f"Row {i+1}: Title is required for classification")
                        failure_count += 1
                        continue
                    
                    target_sheet = self.classify_product_category(title, data)
                    
                    # Get category number using AI
                    category_number, lookup_info = self.get_category_number_for_product(data)
                    if category_number:
                        data['カテゴリ'] = category_number
                        logger.info(f"✅ Added category number {category_number} to product {i+1}")
                    else:
                        logger.warning(f"⚠️ Could not determine category number for product {i+1}: {title}")
                        data['カテゴリ'] = ""  # Leave empty if not found
                    
                    # Map data to sheet headers
                    mapped_data = self.map_data_to_sheet_headers(data, target_sheet)
                    if not mapped_data:
                        error_messages.append(f"Row {i+1}: Failed to map data for sheet: {target_sheet}")
                        failure_count += 1
                        continue
                    
                    # Generate measurement text if applicable
                    measurement_text = self.generate_measurement_text(data, target_sheet)
                    if measurement_text:
                        mapped_data['採寸1'] = measurement_text
                        # Only set 採寸2 if it doesn't already have data
                        if not mapped_data.get('採寸2'):
                            mapped_data['採寸2'] = measurement_text
                    
                    # Check if sheet exists
                    if target_sheet not in wb.sheetnames:
                        error_messages.append(f"Row {i+1}: Sheet '{target_sheet}' not found in workbook")
                        failure_count += 1
                        continue
                    
                    ws = wb[target_sheet]
                    
                    # Get headers from first row
                    headers = [cell.value for cell in ws[1] if cell.value]
                    
                    # Prepare row data in correct order
                    row_data = []
                    for header in headers:
                        value = mapped_data.get(header, '')
                        # Convert None to empty string
                        if value is None:
                            value = ''
                        row_data.append(value)
                    
                    # Find the next empty row
                    next_row = ws.max_row + 1
                    
                    # Write data to the next row
                    for col, value in enumerate(row_data, 1):
                        cell = ws.cell(row=next_row, column=col, value=value)
                        
                        # Add border to data cells
                        thin_border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                        cell.border = thin_border
                    
                    logger.info(f"Data added to {target_sheet} at row {next_row}")
                    success_count += 1
                    
                except Exception as e:
                    failure_count += 1
                    error_messages.append(f"Row {i+1}: Unexpected error - {str(e)}")
                    continue
            
            # Save the workbook
            if success_count > 0:
                logger.info(f"💾 Saving Excel workbook with {success_count} new entries")
                wb.save(output_path)
                logger.info(f"✅ Excel workbook saved successfully")
            
            wb.close()
            
        except Exception as e:
            error_messages.append(f"Critical error during Excel creation: {str(e)}")
            failure_count = len(data_list)  # Mark all as failed if critical error occurs
            success_count = 0
        
        return success_count, failure_count, error_messages
    
    def get_sheet_info(self) -> Dict:
        """
        Get information about all sheets using predefined headers.
        """
        try:
            sheet_info = {}
            
            # Use predefined headers instead of reading from Excel file
            for sheet_name, headers in self.SHEET_HEADERS.items():
                sheet_info[sheet_name] = {
                    'headers': headers,
                    'header_count': len(headers)
                }
            
            return sheet_info
            
        except Exception as e:
            return {'error': f"Could not get sheet info: {str(e)}"}

