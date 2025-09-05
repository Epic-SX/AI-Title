"""
Excel data service for classifying and adding product data to the appropriate sheets
in the PL出品マクロ.xlsm file.
"""

from openpyxl import load_workbook
from typing import Dict, List, Tuple, Optional
import re
import os
import time
import logging

# Set up logging
logger = logging.getLogger(__name__)

class ExcelDataService:
    def __init__(self, excel_file_path: str = None):
        if excel_file_path is None:
            # Use absolute path relative to this file's directory
            current_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
            excel_file_path = os.path.join(current_dir, "PL出品マクロ.xlsm")
        self.excel_file_path = excel_file_path
        logger.info(f"📊 Excel file path: {self.excel_file_path}")
        logger.info(f"📊 Excel file exists: {os.path.exists(self.excel_file_path)}")
        
        # Pre-defined sheet headers (constant structure like your sample code)
        # These headers match exactly what's in the PL出品マクロ.xlsm file
        self.SHEET_HEADERS = {
            "トップス": [
                "カテゴリ", "管理番号", "タイトル", "文字数", "付属品", "ランク", "コメント", 
                "素材", "色", "サイズ", "着丈", "　肩幅", "身幅", "袖丈", "梱包サイズ", 
                "梱包記号", "美品", "ブランド", "フリー", "袖", "もの", "男女", 
                "採寸1", "ラック", "金額", "股上", "股下", "ウエスト", "もも幅", "裾幅", "総丈", "ヒップ", "仕入先", "仕入日", "原価"
            ],
            "パンツ": [
                "カテゴリ", "管理番号", "タイトル", "文字数", "付属品", "ランク", "コメント", 
                "素材", "色", "サイズ", "着丈", "　肩幅", "身幅", "袖丈", "梱包サイズ", 
                "梱包記号", "美品", "ブランド", "フリー", "袖", "もの", "男女", 
                "採寸1", "ラック", "金額", "股上", "股下", "ウエスト", "もも幅", "裾幅", "総丈", "ヒップ", "仕入先", "仕入日", "原価"
            ],
            "スカート": [
                "カテゴリ", "管理番号", "タイトル", "文字数", "付属品", "ランク", "コメント", 
                "素材", "色", "サイズ", "着丈", "　肩幅", "身幅", "袖丈", "梱包サイズ", 
                "梱包記号", "美品", "ブランド", "フリー", "袖", "もの", "男女", 
                "採寸1", "ラック", "金額", "股上", "股下", "ウエスト", "もも幅", "裾幅", "総丈", "ヒップ", "仕入先", "仕入日", "原価"
            ],
            "ワンピース": [
                "カテゴリ", "管理番号", "タイトル", "文字数", "付属品", "ランク", "コメント", 
                "素材", "色", "サイズ", "着丈", "　肩幅", "身幅", "袖丈", "梱包サイズ", 
                "梱包記号", "美品", "ブランド", "フリー", "袖", "もの", "男女", 
                "採寸1", "ラック", "金額", "股上", "股下", "ウエスト", "もも幅", "裾幅", "総丈", "ヒップ", "仕入先", "仕入日", "原価"
            ],
            "オールインワン": [
                "カテゴリ", "管理番号", "タイトル", "文字数", "付属品", "ランク", "コメント", 
                "素材", "色", "サイズ", "着丈", "　肩幅", "身幅", "袖丈", "梱包サイズ", 
                "梱包記号", "美品", "ブランド", "フリー", "袖", "もの", "男女", 
                "採寸1", "ラック", "金額", "股上", "股下", "ウエスト", "もも幅", "裾幅", "総丈", "ヒップ", "仕入先", "仕入日", "原価"
            ],
            "スカートスーツ": [
                "カテゴリ", "管理番号", "タイトル", "文字数", "付属品", "ランク", "コメント", 
                "素材", "色", "サイズ", "着丈", "　肩幅", "身幅", "袖丈", "梱包サイズ", 
                "梱包記号", "美品", "ブランド", "フリー", "袖", "もの", "男女", 
                "採寸1", "ラック", "金額", "股上", "股下", "ウエスト", "もも幅", "裾幅", "総丈", "ヒップ", "仕入先", "仕入日", "原価"
            ],
            "パンツスーツ": [
                "カテゴリ", "管理番号", "タイトル", "文字数", "付属品", "ランク", "コメント", 
                "素材", "色", "サイズ", "着丈", "　肩幅", "身幅", "袖丈", "梱包サイズ", 
                "梱包記号", "美品", "ブランド", "フリー", "袖", "もの", "男女", 
                "採寸1", "ラック", "金額", "股上", "股下", "ウエスト", "もも幅", "裾幅", "総丈", "ヒップ", "仕入先", "仕入日", "原価"
            ],
            "アンサンブル": [
                "カテゴリ", "管理番号", "タイトル", "文字数", "付属品", "ランク", "コメント", 
                "素材", "色", "サイズ", "着丈", "　肩幅", "身幅", "袖丈", "梱包サイズ", 
                "梱包記号", "美品", "ブランド", "フリー", "袖", "もの", "男女", 
                "採寸1", "ラック", "金額", "股上", "股下", "ウエスト", "もも幅", "裾幅", "総丈", "ヒップ", "仕入先", "仕入日", "原価"
            ],
            "靴": [
                "カテゴリ", "管理番号", "タイトル", "文字数", "付属品", "ランク", "コメント", 
                "素材", "色", "サイズ", "着丈", "　肩幅", "身幅", "袖丈", "梱包サイズ", 
                "梱包記号", "美品", "ブランド", "フリー", "袖", "もの", "男女", 
                "採寸1", "ラック", "金額", "股上", "股下", "ウエスト", "もも幅", "裾幅", "総丈", "ヒップ", "仕入先", "仕入日", "原価"
            ],
            "ブーツ": [
                "カテゴリ", "管理番号", "タイトル", "文字数", "付属品", "ランク", "コメント", 
                "素材", "色", "サイズ", "着丈", "　肩幅", "身幅", "袖丈", "梱包サイズ", 
                "梱包記号", "美品", "ブランド", "フリー", "袖", "もの", "男女", 
                "採寸1", "ラック", "金額", "股上", "股下", "ウエスト", "もも幅", "裾幅", "総丈", "ヒップ", "仕入先", "仕入日", "原価"
            ],
            "ベルト": [
                "カテゴリ", "管理番号", "タイトル", "文字数", "付属品", "ランク", "コメント", 
                "素材", "色", "サイズ", "着丈", "　肩幅", "身幅", "袖丈", "梱包サイズ", 
                "梱包記号", "美品", "ブランド", "フリー", "袖", "もの", "男女", 
                "採寸1", "ラック", "金額", "股上", "股下", "ウエスト", "もも幅", "裾幅", "総丈", "ヒップ", "仕入先", "仕入日", "原価"
            ],
            "ネクタイ縦横": [
                "カテゴリ", "管理番号", "タイトル", "文字数", "付属品", "ランク", "コメント", 
                "素材", "色", "サイズ", "着丈", "　肩幅", "身幅", "袖丈", "梱包サイズ", 
                "梱包記号", "美品", "ブランド", "フリー", "袖", "もの", "男女", 
                "採寸1", "ラック", "金額", "股上", "股下", "ウエスト", "もも幅", "裾幅", "総丈", "ヒップ", "仕入先", "仕入日", "原価"
            ],
            "帽子": [
                "カテゴリ", "管理番号", "タイトル", "文字数", "付属品", "ランク", "コメント", 
                "素材", "色", "サイズ", "着丈", "　肩幅", "身幅", "袖丈", "梱包サイズ", 
                "梱包記号", "美品", "ブランド", "フリー", "袖", "もの", "男女", 
                "採寸1", "ラック", "金額", "股上", "股下", "ウエスト", "もも幅", "裾幅", "総丈", "ヒップ", "仕入先", "仕入日", "原価"
            ],
            "バッグ": [
                "カテゴリ", "管理番号", "タイトル", "文字数", "付属品", "ランク", "コメント", 
                "素材", "色", "サイズ", "着丈", "　肩幅", "身幅", "袖丈", "梱包サイズ", 
                "梱包記号", "美品", "ブランド", "フリー", "袖", "もの", "男女", 
                "採寸1", "ラック", "金額", "股上", "股下", "ウエスト", "もも幅", "裾幅", "総丈", "ヒップ", "仕入先", "仕入日", "原価"
            ],
            "ネックレス": [
                "カテゴリ", "管理番号", "タイトル", "文字数", "付属品", "ランク", "コメント", 
                "素材", "色", "サイズ", "着丈", "　肩幅", "身幅", "袖丈", "梱包サイズ", 
                "梱包記号", "美品", "ブランド", "フリー", "袖", "もの", "男女", 
                "採寸1", "ラック", "金額", "股上", "股下", "ウエスト", "もも幅", "裾幅", "総丈", "ヒップ", "仕入先", "仕入日", "原価"
            ],
            "サングラス": [
                "カテゴリ", "管理番号", "タイトル", "文字数", "付属品", "ランク", "コメント", 
                "素材", "色", "サイズ", "着丈", "　肩幅", "身幅", "袖丈", "梱包サイズ", 
                "梱包記号", "美品", "ブランド", "フリー", "袖", "もの", "男女", 
                "採寸1", "ラック", "金額", "股上", "股下", "ウエスト", "もも幅", "裾幅", "総丈", "ヒップ", "仕入先", "仕入日", "原価"
            ]
        }
        
        # Category classification keywords for fallback
        self.category_keywords = {
            'トップス': [
                'ブラウス', 'シャツ', 'tシャツ', 'カットソー', 'ニット', 'セーター', 
                'パーカー', 'フリース', 'ジャケット', 'カーディガン', 'ベスト',
                'タンクトップ', 'キャミソール', 'チュニック'
            ],
            'パンツ': [
                'パンツ', 'ズボン', 'ジーンズ', 'デニム', 'チノパン', 'スラックス',
                'レギンス', 'ショートパンツ', 'ハーフパンツ', 'ワイドパンツ',
                'スキニー', 'ボトムス', 'トラウザー'
            ],
            'スカート': [
                'スカート', 'ミニスカート', 'ロングスカート', 'マキシスカート',
                'フレアスカート', 'タイトスカート', 'プリーツスカート'
            ],
            'ワンピース': [
                'ワンピース', 'ドレス', 'マキシワンピース', 'ミニワンピース',
                'シャツワンピース', 'ニットワンピース'
            ],
            'オールインワン': [
                'オールインワン', 'サロペット', 'オーバーオール', 'ジャンプスーツ',
                'コンビネゾン', 'つなぎ'
            ],
            'スカートスーツ': [
                'スカートスーツ', 'スーツ.*スカート', 'セットアップ.*スカート'
            ],
            'パンツスーツ': [
                'パンツスーツ', 'スーツ.*パンツ', 'セットアップ.*パンツ'
            ],
            'アンサンブル': [
                'アンサンブル', 'ツインセット', 'セット.*ニット'
            ],
            '靴': [
                'パンプス', 'ヒール', 'フラットシューズ', '革靴', 'ローファー',
                'サンダル', 'ミュール', 'オックスフォード'
            ],
            'ブーツ': [
                'ブーツ', 'ロングブーツ', 'ショートブーツ', 'アンクルブーツ',
                'ニーハイブーツ', 'ムートンブーツ'
            ],
            'ベルト': [
                'ベルト', 'レザーベルト', 'チェーンベルト'
            ],
            'ネクタイ縦横': [
                'ネクタイ', 'タイ', 'ボウタイ'
            ],
            '帽子': [
                '帽子', 'キャップ', 'ハット', 'ベレー帽', 'ニット帽',
                'ビーニー', '麦わら帽子', 'ハンチング'
            ],
            'バッグ': [
                'バッグ', 'ハンドバッグ', 'ショルダーバッグ', 'トートバッグ',
                'クラッチバッグ', 'リュック', 'バックパック', 'ポーチ',
                'ウエストバッグ', 'メッセンジャーバッグ'
            ],
            'ネックレス': [
                'ネックレス', 'チョーカー', 'ペンダント', 'チェーン'
            ],
            'サングラス': [
                'サングラス', 'メガネ', '眼鏡', 'グラス'
            ]
        }
        
        # Default sheet for unclassified items
        self.default_sheet = 'トップス'
        
    def classify_product_category(self, title: str, product_data: Dict) -> str:
        """
        Classify product category based on title and product data.
        Returns the appropriate sheet name.
        """
        title_lower = title.lower()
        
        # Check if product_data has specific type information
        product_type = product_data.get('もの', '') or product_data.get('product_type', '')
        if product_type:
            title_lower += ' ' + product_type.lower()
        
        # Check for keywords in each category
        for category, keywords in self.category_keywords.items():
            for keyword in keywords:
                if re.search(keyword.lower(), title_lower):
                    return category
        
        # If no specific category found, return default
        return self.default_sheet
    
    def map_data_to_sheet_headers(self, data: Dict, sheet_name: str) -> Dict:
        """
        Map the input data to the appropriate headers for the target sheet.
        Uses predefined headers instead of reading from Excel file.
        """
        try:
            # Use predefined headers instead of reading from Excel file
            if sheet_name not in self.SHEET_HEADERS:
                logger.error(f"Sheet '{sheet_name}' not found in predefined headers")
                return {}
            
            headers = self.SHEET_HEADERS[sheet_name]
            
        except Exception as e:
            logger.error(f"Error getting sheet headers: {e}")
            return {}
        
        # Create mapping based on common field names
        field_mappings = {
            'カテゴリ': ['category', 'カテゴリ'],
            '管理番号': ['management_number', '管理番号', 'id'],
            'タイトル': ['title', 'タイトル'],
            '文字数': ['character_count', '文字数'],
            '付属品': ['accessories', '付属品'],
            '日本サイズ': ['japanese_size', '日本サイズ'],
            'ランク': ['rank', 'ランク', 'condition_rank'],
            'コメント': ['comment', 'コメント', 'description'],
            '仕立て・収納': ['tailoring_storage', '仕立て・収納'],
            '素材': ['material', '素材'],
            '色': ['color', '色'],
            'サイズ': ['size', 'サイズ'],
            '梱包サイズ': ['packaging_size', '梱包サイズ'],
            '梱包記号': ['packaging_symbol', '梱包記号'],
            '美品': ['excellent_condition', '美品'],
            'ブランド': ['brand', 'ブランド'],
            'フリー': ['free_text', 'フリー'],
            '袖': ['sleeve', '袖'],
            'もの': ['item_type', 'もの', 'product_type'],
            '男女': ['gender', '男女'],
            'ラック': ['rack', 'ラック'],
            '仕入先': ['supplier', '仕入先'],
            '仕入日': ['purchase_date', '仕入日'],
            '原価': ['cost_price', '原価'],
            '金額': ['price', '金額', 'amount'],
            # Measurement fields
            '着丈': ['garment_length', '着丈'],
            '　肩幅': ['shoulder_width', 'shoulder_width', '肩幅', '　肩幅'],
            '身幅': ['chest_width', '身幅'],
            '袖丈': ['sleeve_length', '袖丈'],
            '股上': ['rise', '股上'],
            '股下': ['inseam', '股下'],
            'ウエスト': ['waist', 'ウエスト'],
            'もも幅': ['thigh_width', 'もも幅'],
            '裾幅': ['hem_width', '裾幅'],
            '総丈': ['total_length', '総丈'],
            'ヒップ': ['hip', 'ヒップ'],
            '採寸1': ['measurement1', '採寸1'],
            '採寸2': ['measurement2', '採寸2']
        }
        
        # Create the mapped data
        mapped_data = {}
        
        for header in headers:
            # Find matching field in input data
            value = None
            
            # Direct match
            if header in data:
                value = data[header]
            else:
                # Check mappings
                if header in field_mappings:
                    for field_key in field_mappings[header]:
                        if field_key in data:
                            value = data[field_key]
                            break
            
            # Set the value (None will be handled as empty)
            mapped_data[header] = value
        
        return mapped_data
    
    def generate_measurement_text(self, data: Dict, sheet_name: str) -> str:
        """
        Generate formatted measurement text based on the product category.
        """
        measurements = []
        
        if sheet_name == 'トップス':
            if data.get('着丈'):
                measurements.append(f"着丈：約{data['着丈']}cm")
            if data.get('　肩幅') or data.get('肩幅'):
                shoulder = data.get('　肩幅') or data.get('肩幅')
                measurements.append(f"肩幅：約{shoulder}cm")
            if data.get('身幅'):
                measurements.append(f"身幅：約{data['身幅']}cm")
            if data.get('袖丈'):
                measurements.append(f"袖丈：約{data['袖丈']}cm")
                
        elif sheet_name == 'パンツ':
            if data.get('股上'):
                measurements.append(f"股上：約{data['股上']}cm")
            if data.get('股下'):
                measurements.append(f"股下：約{data['股下']}cm")
            if data.get('ウエスト'):
                measurements.append(f"ウエスト：約{data['ウエスト']}cm")
            if data.get('もも幅'):
                measurements.append(f"もも幅：約{data['もも幅']}cm")
            if data.get('裾幅'):
                measurements.append(f"裾幅：約{data['裾幅']}cm")
                
        elif sheet_name == 'スカート':
            if data.get('総丈'):
                measurements.append(f"総丈：約{data['総丈']}cm")
            if data.get('ウエスト'):
                measurements.append(f"ウエスト：約{data['ウエスト']}cm")
            if data.get('ヒップ'):
                measurements.append(f"ヒップ：約{data['ヒップ']}cm")
        
        return "　".join(measurements) if measurements else ""
    
    def add_data_to_excel(self, data: Dict) -> Tuple[bool, str]:
        """
        Add product data to the appropriate sheet in the Excel file.
        
        Args:
            data: Dictionary containing product data
            
        Returns:
            Tuple of (success: bool, message: str)
        """
        try:
            # Check if file exists
            if not os.path.exists(self.excel_file_path):
                return False, f"Excel file not found: {self.excel_file_path}"
            
            # Classify the product to determine target sheet
            title = data.get('タイトル', '') or data.get('title', '')
            if not title:
                return False, "Title is required for classification"
            
            target_sheet = self.classify_product_category(title, data)
            
            # Map data to sheet headers
            mapped_data = self.map_data_to_sheet_headers(data, target_sheet)
            if not mapped_data:
                return False, f"Failed to map data for sheet: {target_sheet}"
            
            # Generate measurement text if applicable
            measurement_text = self.generate_measurement_text(data, target_sheet)
            if measurement_text:
                mapped_data['採寸1'] = measurement_text
                # Only set 採寸2 if it doesn't already have data
                if not mapped_data.get('採寸2'):
                    mapped_data['採寸2'] = measurement_text
            
            # Load workbook with macros preserved
            book = load_workbook(self.excel_file_path, keep_vba=True)
            
            if target_sheet not in book.sheetnames:
                return False, f"Sheet '{target_sheet}' not found in workbook"
            
            ws = book[target_sheet]
            
            # Get headers from first row
            headers = [cell.value for cell in ws[1] if cell.value]
            
            # Prepare row data in correct order
            row_data = []
            for header in headers:
                value = mapped_data.get(header, '')
                # Convert None to empty string
                if value is None:
                    value = ''
                row_data.append(value)
            
            # Function to check if a row is empty (all cells are None or empty string)
            def is_row_empty(row_num, num_columns):
                for col in range(1, num_columns + 1):
                    cell_value = ws.cell(row=row_num, column=col).value
                    if cell_value is not None and str(cell_value).strip() != "":
                        return False
                return True
            
            # Find the first empty row
            num_columns = len(row_data)
            target_row = None
            
            # Start from row 2 (assuming row 1 has headers)
            for row in range(2, ws.max_row + 1):
                if is_row_empty(row, num_columns):
                    target_row = row
                    break
            
            # If no empty row found, use the next row after the last data
            if target_row is None:
                target_row = ws.max_row + 1
            
            logger.info(f"Found empty row at: {target_row}")
            
            # Write data to the target row
            for col, value in enumerate(row_data, 1):
                ws.cell(row=target_row, column=col, value=value)
            
            logger.info(f"Data added to row {target_row}")
            
            # Save the workbook
            book.save(self.excel_file_path)
            
            return True, f"Data successfully added to sheet: {target_sheet} at row {target_row}"
            
        except Exception as e:
            return False, f"Error adding data to Excel: {str(e)}"
    
    def get_sheet_info(self) -> Dict:
        """
        Get information about all sheets using predefined headers.
        """
        try:
            sheet_info = {}
            
            # Use predefined headers instead of reading from Excel file
            for sheet_name, headers in self.SHEET_HEADERS.items():
                sheet_info[sheet_name] = {
                    'headers': headers,
                    'header_count': len(headers)
                }
            
            return sheet_info
            
        except Exception as e:
            return {'error': f"Could not get sheet info: {str(e)}"}
    
    def bulk_add_data(self, data_list: List[Dict]) -> Tuple[int, int, List[str]]:
        """
        Add multiple product data entries to the Excel file efficiently.
        Loads the workbook once and saves it once at the end.
        
        Args:
            data_list: List of dictionaries containing product data
            
        Returns:
            Tuple of (success_count: int, failure_count: int, error_messages: List[str])
        """
        if not data_list:
            return 0, 0, []
        
        success_count = 0
        failure_count = 0
        error_messages = []
        
        try:
            # Load workbook once for all operations
            logger.info(f"📊 Loading Excel workbook for bulk operation: {self.excel_file_path}")
            book = load_workbook(self.excel_file_path, keep_vba=True)
            
            # Process all products
            for i, data in enumerate(data_list):
                try:
                    # Classify the product to determine target sheet
                    title = data.get('タイトル', '') or data.get('title', '')
                    if not title:
                        error_messages.append(f"Row {i+1}: Title is required for classification")
                        failure_count += 1
                        continue
                    
                    target_sheet = self.classify_product_category(title, data)
                    
                    # Map data to sheet headers
                    mapped_data = self.map_data_to_sheet_headers(data, target_sheet)
                    if not mapped_data:
                        error_messages.append(f"Row {i+1}: Failed to map data for sheet: {target_sheet}")
                        failure_count += 1
                        continue
                    
                    # Generate measurement text if applicable
                    measurement_text = self.generate_measurement_text(data, target_sheet)
                    if measurement_text:
                        mapped_data['採寸1'] = measurement_text
                        # Only set 採寸2 if it doesn't already have data
                        if not mapped_data.get('採寸2'):
                            mapped_data['採寸2'] = measurement_text
                    
                    # Check if sheet exists
                    if target_sheet not in book.sheetnames:
                        error_messages.append(f"Row {i+1}: Sheet '{target_sheet}' not found in workbook")
                        failure_count += 1
                        continue
                    
                    ws = book[target_sheet]
                    
                    # Get headers from first row
                    headers = [cell.value for cell in ws[1] if cell.value]
                    
                    # Prepare row data in correct order
                    row_data = []
                    for header in headers:
                        value = mapped_data.get(header, '')
                        # Convert None to empty string
                        if value is None:
                            value = ''
                        row_data.append(value)
                    
                    # Function to check if a row is empty (all cells are None or empty string)
                    def is_row_empty(row_num, num_columns):
                        for col in range(1, num_columns + 1):
                            cell_value = ws.cell(row=row_num, column=col).value
                            if cell_value is not None and str(cell_value).strip() != "":
                                return False
                        return True
                    
                    # Find the first empty row
                    num_columns = len(row_data)
                    target_row = None
                    
                    # Start from row 2 (assuming row 1 has headers)
                    for row in range(2, ws.max_row + 1):
                        if is_row_empty(row, num_columns):
                            target_row = row
                            break
                    
                    # If no empty row found, use the next row after the last data
                    if target_row is None:
                        target_row = ws.max_row + 1
                    
                    logger.info(f"Found empty row at: {target_row}")
                    
                    # Write data to the target row
                    for col, value in enumerate(row_data, 1):
                        ws.cell(row=target_row, column=col, value=value)
                    
                    logger.info(f"Data added to row {target_row}")
                    success_count += 1
                    
                except Exception as e:
                    failure_count += 1
                    error_messages.append(f"Row {i+1}: Unexpected error - {str(e)}")
                    continue
            
            # Save the workbook once after all operations
            if success_count > 0:
                logger.info(f"💾 Saving Excel workbook with {success_count} new entries")
                book.save(self.excel_file_path)
                logger.info(f"✅ Excel workbook saved successfully")
            
        except Exception as e:
            error_messages.append(f"Critical error during bulk operation: {str(e)}")
            failure_count += len(data_list)  # Mark all as failed if critical error occurs
            success_count = 0
        
        return success_count, failure_count, error_messages 